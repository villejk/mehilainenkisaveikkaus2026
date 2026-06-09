"""
update_data.py
--------------
Lukee MM 2026 -veikkausExcelin SharePointista ja generoi data.json-tiedoston
GitHub-repositorioon. Ajetaan GitHub Actionsissa automaattisesti.

Ympäristömuuttujat (GitHub Secrets):
  SHAREPOINT_SITE_URL   esim. https://mehilainen.sharepoint.com/sites/yoursite
  SHAREPOINT_FILE_PATH  esim. /sites/yoursite/Shared Documents/MM2026/tiedosto.xlsm
  AZURE_CLIENT_ID       Azure App Registration client ID
  AZURE_CLIENT_SECRET   Azure App Registration client secret
  AZURE_TENANT_ID       Azure tenant ID
"""

import os
import json
import sys
import io
import tempfile
import requests
from datetime import datetime, timezone
from openpyxl import load_workbook


# ── Konfiguraatio ──────────────────────────────────────────────────────────────

SHAREPOINT_SITE_URL  = os.environ["SHAREPOINT_SITE_URL"]   # https://org.sharepoint.com/sites/xxx
SHAREPOINT_FILE_PATH = os.environ["SHAREPOINT_FILE_PATH"]  # /sites/xxx/Shared Documents/.../file.xlsm
AZURE_CLIENT_ID      = os.environ["AZURE_CLIENT_ID"]
AZURE_CLIENT_SECRET  = os.environ["AZURE_CLIENT_SECRET"]
AZURE_TENANT_ID      = os.environ["AZURE_TENANT_ID"]

OUTPUT_PATH = "data.json"


# ── 1. Hae access token Microsoft Graphiin ────────────────────────────────────

def get_access_token() -> str:
    url = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}/oauth2/v2.0/token"
    resp = requests.post(url, data={
        "grant_type":    "client_credentials",
        "client_id":     AZURE_CLIENT_ID,
        "client_secret": AZURE_CLIENT_SECRET,
        "scope":         "https://graph.microsoft.com/.default",
    })
    resp.raise_for_status()
    return resp.json()["access_token"]


# ── 2. Lataa Excel SharePointista ────────────────────────────────────────────

def download_excel(token: str) -> bytes:
    """Lataa Excel-tiedosto SharePointista Graph API:n kautta."""
    # Muodosta Graph API -URL tiedostolle
    # SHAREPOINT_FILE_PATH muodossa /sites/xxx/Shared Documents/Kansio/tiedosto.xlsm
    encoded = requests.utils.quote(SHAREPOINT_FILE_PATH)
    graph_url = f"https://graph.microsoft.com/v1.0/sites/root:{encoded}:/content"

    headers = {"Authorization": f"Bearer {token}"}
    resp = requests.get(graph_url, headers=headers)

    if resp.status_code != 200:
        # Vaihtoehtoinen tapa: hae site ID:llä
        site_hostname = SHAREPOINT_SITE_URL.replace("https://", "").split("/")[0]
        site_path = "/" + "/".join(SHAREPOINT_SITE_URL.replace("https://", "").split("/")[1:])
        site_resp = requests.get(
            f"https://graph.microsoft.com/v1.0/sites/{site_hostname}:{site_path}",
            headers=headers
        )
        site_resp.raise_for_status()
        site_id = site_resp.json()["id"]

        # Hae tiedosto site ID:llä
        file_resp = requests.get(
            f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:{SHAREPOINT_FILE_PATH}:/content",
            headers=headers
        )
        file_resp.raise_for_status()
        return file_resp.content

    return resp.content


# ── 3. Parseri: Pistetilanne-välilehti ───────────────────────────────────────

def parse_pistetilanne(wb) -> dict:
    ws = wb["Pistetilanne"]
    rows = list(ws.iter_rows(values_only=True))

    players = []
    for row in rows[3:]:
        # Pelaajarivin tunnistus: sija on numero, nimi on merkkijono >1 merkkiä
        if (row[0] is not None and row[1] is not None
                and isinstance(row[1], str) and len(row[1]) > 1
                and isinstance(row[0], (int, float))):
            players.append({
                "sija":          int(row[0]),
                "pelaaja":       row[1],
                "pisteet":       float(row[2] or 0),
                "osa_a":         float(row[3] or 0),
                "osa_b":         float(row[4] or 0),
                "osa_c":         float(row[5] or 0),
                "osa_d":         float(row[6] or 0),
                "osa_e":         float(row[7] or 0),
                "osa_f":         float(row[8] or 0),
                "osa_g":         float(row[9] or 0),
                "voittaja":      row[10] or "-",
                "korttikuningas":row[11] or "-",
                "maalikuningas": row[12] or "-",
            })

    # Laske sijoitukset uudelleen pisteiden mukaan
    players.sort(key=lambda p: -p["pisteet"])
    prev_pts, rank = None, 0
    for i, p in enumerate(players):
        if p["pisteet"] != prev_pts:
            rank = i + 1
        p["sija"] = rank
        prev_pts = p["pisteet"]

    return players


# ── 4. Parseri: Pistekertymä-välilehti ───────────────────────────────────────

def parse_pistekertymä(wb) -> dict:
    ws = wb["Pistekertymä"]
    rows = list(ws.iter_rows(values_only=True))

    # Rivi 2 (indeksi 2): otsikkorivi jossa pelaajaanimet sarakkeissa 4→
    header = rows[2]
    player_names = [h for h in header[4:] if h]

    matches = []
    for row in rows[3:]:
        if row[1] is None:
            continue
        aika = row[2].strftime("%d.%m. %H:%M") if hasattr(row[2], "strftime") else str(row[2] or "")
        match = {
            "nro":    int(row[1]),
            "aika":   aika,
            "ottelu": row[3] or "",
        }
        for i, name in enumerate(player_names):
            val = row[4 + i]
            match[name] = float(val) if val is not None else 0.0
        matches.append(match)

    return matches, player_names


# ── 5. Kirjoita data.json ─────────────────────────────────────────────────────

def build_json(players, matches, player_names) -> dict:
    return {
        "updated":      datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "player_names": player_names,
        "players":      players,
        "matches":      matches,
    }


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("🔐 Haetaan access token...")
    token = get_access_token()

    print("📥 Ladataan Excel SharePointista...")
    excel_bytes = download_excel(token)

    print("📊 Parsitaan Excel...")
    wb = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)

    players              = parse_pistetilanne(wb)
    matches, player_names = parse_pistekertymä(wb)

    data = build_json(players, matches, player_names)

    print(f"✅ {len(players)} pelaajaa, {len(matches)} ottelua")
    print(f"   Johtaja: {players[0]['pelaaja']} ({players[0]['pisteet']} p)" if players else "")

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"💾 Kirjoitettu → {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
